'''!
This script is the entry point to the sherpa tool.

Call it as:

python main.py ./kewords.json

where keywords.json contains a list of keywords and the markets for the SERP request. E.g.:

[
    {
        "name": "keyword1",
        "keywords": ["emergency notification system",
                    "enterprise mass notification",
                    "critical event management"]
        "markets": ["AT", "BE", "CH", "CY", "CZ", "DE", "DK", "ES", "FI", "FR", "GB", "GR", "HR", "IE", "IT", "NL", "NO", "PL",
                    "PT", "SE", "SI", "SK"]
    },
    {
        "name": "keyword1_es",
        "keywords": ["sistema de notificaciones masivas",
                    "solución de alertas corporativas",
                    "sistema de notificaciones corporativas",
                    "gestión eventos críticos"]
        "markets": ["es"]
    }
]

'''

import sys
from mlogger import logger
from pathlib import Path
from apis import GoogleSearchApi
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import json

logger = logger.get_module_logger(__name__)

def usage():
    print("Usage: python main.py keywords.json")
    print("		keywords.json 	JSON file with keyword objects")
    sys.exit(1)

def parse_keys_file():
    '''
    Parses the secrets.json file to retrieve all the keys for the different APIs.
    @return: Object that contains the list of API keys.
    '''
    logger.info("Parsing lists of API keys.")
    # The key is the token "api" field in each token inside tokens.json
    keyspath = Path.cwd() / "secrets/api_keys.json"
    assert (keyspath.exists())
    with open(keyspath) as keysfile:
        keys = json.load(keysfile)
    logger.debug(f"API keys: {keys}")
    return keys

# def storeResult(result):
#     '''
#     Generates one CVS file per keyword in results.
#     The output CVS files are named using the keyword name suffixed with the current date.
#     :param result: SERP results
#     :return: void
#     '''
#     for key in result.keys():
#         filename = key + "_" + date.today().strftime("%d-%m-%Y") + ".csv"
#         header = ",".join(result[key].keys())  # list of markets
#         with open(filename, "w") as outfile:
#             outfile.write(header + "\n")
#             longest_res_length = 0  # tracks the market with the largest number of results
#             # first pass to find the largets length
#             for market in result[key]:
#                 longest_res_length = len(result[key][market]) if (len(result[key][market]) > longest_res_length) else longest_res_length
#             # second pass is to build the lines of the CSV file
#             for i in range(longest_res_length):
#                 line = ""
#                 for market in result[key]: # convert lines to columns
#                     if len(result[key][market]) < i:
#                         line = line + ","  # insert empty line
#                     else:
#                         line = line + result[key][market][i] + ","
#                 outfile.write(line[:len(line) - 1] + "\n")    # store the line without the last comma

def storeResult(result, outfilename = None):
    '''
    Generates one CVS file per keyword in results.
    The output CVS files are named using the keyword name suffixed with the current date.
    :param result: SERP results
    :param outfilename: name of the output excel file without sufix
    :return: void
    '''
    if outfilename == None:
        outfilename = "sherpa_" + date.today().strftime("%d-%m-%Y")
    wb = Workbook()
    for key in result.keys():
        sheets = wb.sheetnames
        sheetname = key
        if sheetname in sheets:
            sheetname += sheetname
        ws = wb.create_sheet(sheetname)
        # Format
        a1 = ws['A1']
        b1 = ws['B1']
        ft = Font(size=12, bold=True)
        a1.font = ft
        b1.font = ft
        ws.append(["Keyword", key])
        logger.debug(f"Adding header to worksheet: {result[key].keys()}")
        ws.append(list(result[key].keys()))
        for i in range(1, len(result[key].keys()) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 50
        longest_res_length = 0  # tracks the market with the largest number of results
        # first pass to find the largets length
        for market in result[key]:
            longest_res_length = len(result[key][market]) if (len(result[key][market]) > longest_res_length) else longest_res_length
        # second pass is to build the lines of the CSV file
        for i in range(longest_res_length):
            line = []
            for market in result[key]: # convert lines to columns
                if len(result[key][market]) < i:
                    line.append("")  # insert empty line
                else:
                    line.append(result[key][market][i])
            # logger.debug(f"Appending {line}")
            ws.append(line)    # store the line
    wb.remove(wb[wb.sheetnames[0]])
    wb.save(outfilename + ".xlsx")

if __name__ == '__main__':
    logger.info("Starting SERP queries.")
    api_keys = parse_keys_file()
    if(len(sys.argv) < 2):
        usage()
    keywordpath = Path(sys.argv[1])
    assert (keywordpath.exists())
    with open(keywordpath) as keywordfile:
        keywords = json.load(keywordfile)
        gs_api = GoogleSearchApi(api_keys["rapidapi"])
        res = {} # will contain an entry for each keyword
        for entry in keywords:
            assert("keyword" in entry.keys())
            assert("markets" in entry.keys())
            logger.debug(f"entry: {entry}")
            res[entry["name"]] = {}
            for market in entry["markets"]:
                keyword = {
                    "keyword": entry["keyword"],
                    "country": market
                }
                resmar = gs_api.queryKeyword(keyword)
                res[entry["name"]][market] = resmar
        # logger.debug(f"Result = {res}")
        storeResult(res)



